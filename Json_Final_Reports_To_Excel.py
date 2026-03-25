import argparse
import json
import os
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def safe_sheet_name(name: str, fallback: str = "Sheet") -> str:
    """
    Excel sheet names:
    - max 31 chars
    - cannot contain: : \\ / ? * [ ]
    """
    name = (name or fallback).strip()
    name = re.sub(r"[\[\]\:\*\?\/\\]", "_", name)
    if not name:
        name = fallback
    if len(name) > 31:
        name = name[:28] + "..."
    return name


def stringify_scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return str(value)
    # For any unexpected type, preserve as JSON.
    return json.dumps(value, ensure_ascii=False)


def iter_leaf_values(obj: Any, path: str = "") -> Iterable[Tuple[str, Any]]:
    """
    Yield (json_path, value) for every "leaf" value in a JSON object.

    Notes:
    - Lists are indexed (e.g. "issues[0].recommendation.fix[1]").
    - Empty dict/list are emitted as leaves to avoid losing structure.
    """
    if isinstance(obj, dict):
        if not obj:
            yield path, {}
            return
        for k, v in obj.items():
            next_path = f"{path}.{k}" if path else str(k)
            yield from iter_leaf_values(v, next_path)
    elif isinstance(obj, list):
        if not obj:
            yield path, []
            return
        for i, v in enumerate(obj):
            next_path = f"{path}[{i}]"
            yield from iter_leaf_values(v, next_path)
    else:
        yield path, obj


def build_all_fields_rows(report: Dict[str, Any]) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for p, v in iter_leaf_values(report):
        if p == "":
            p = "$"  # root
        rows.append({"json_path": p, "value": stringify_scalar(v)})
    return rows


def extract_issues_table(report: Dict[str, Any]) -> Tuple[List[str], List[Dict[str, Any]]]:
    issues = report.get("issues", [])
    if not isinstance(issues, list) or not issues:
        headers = [
                "Issue_Index",
                "rule",
                "category",
                "status",
                "wcag",
                "wcag_name",
                "wcag_level",
                "impact",
                "recommendation.issue",
                "recommendation.priority",
                "recommendation.automation_possible",
                "recommendation.fix",
                "extra_issue_fields_json",
            ]
        return headers, []

    known_issue_keys = {
        "rule",
        "category",
        "status",
        "wcag",
        "wcag_name",
        "wcag_level",
        "impact",
        "recommendation",
    }
    known_rec_keys = {"issue", "fix", "priority", "automation_possible"}

    headers = [
        "Issue_Index",
        "rule",
        "category",
        "status",
        "wcag",
        "wcag_name",
        "wcag_level",
        "impact",
        "recommendation.issue",
        "recommendation.priority",
        "recommendation.automation_possible",
        "recommendation.fix",
        "extra_issue_fields_json",
    ]

    rows: List[Dict[str, Any]] = []
    for idx, item in enumerate(issues):
        if not isinstance(item, dict):
            rows.append({"Issue_Index": idx, "extra_issue_fields_json": stringify_scalar(item)})
            continue

        rec = item.get("recommendation", {}) if isinstance(item.get("recommendation", {}), dict) else {}

        fix_val = rec.get("fix", [])
        if isinstance(fix_val, list):
            fix_text = "\n".join(stringify_scalar(x) for x in fix_val)
        else:
            fix_text = stringify_scalar(fix_val)

        extra_issue = {k: v for k, v in item.items() if k not in known_issue_keys}
        extra_rec = {k: v for k, v in rec.items() if k not in known_rec_keys}

        extra_issue_fields = {}
        if extra_issue:
            extra_issue_fields["extra_issue_fields"] = extra_issue
        if extra_rec:
            extra_issue_fields["extra_recommendation_fields"] = extra_rec

        rows.append(
            {
                "Issue_Index": idx,
                "rule": item.get("rule", ""),
                "category": item.get("category", ""),
                "status": item.get("status", ""),
                "wcag": item.get("wcag", ""),
                "wcag_name": item.get("wcag_name", ""),
                "wcag_level": item.get("wcag_level", ""),
                "impact": item.get("impact", ""),
                "recommendation.issue": rec.get("issue", ""),
                "recommendation.priority": rec.get("priority", ""),
                "recommendation.automation_possible": rec.get("automation_possible", ""),
                "recommendation.fix": fix_text,
                "extra_issue_fields_json": json.dumps(extra_issue_fields, ensure_ascii=False),
            }
        )
    return headers, rows


def extract_manual_review_table(report: Dict[str, Any]) -> Tuple[List[str], List[Dict[str, Any]]]:
    mrr = report.get("manual_review_required", [])
    if not isinstance(mrr, list) or not mrr:
        headers = [
                "Manual_Index",
                "item_type",
                "rule",
                "category",
                "status",
                "wcag",
                "wcag_name",
                "wcag_level",
                "impact",
                "recommendation.issue",
                "recommendation.priority",
                "recommendation.automation_possible",
                "recommendation.fix",
                "value",
                "extra_manual_fields_json",
            ]
        return headers, []

    known_keys = {
        "rule",
        "category",
        "status",
        "wcag",
        "wcag_name",
        "wcag_level",
        "impact",
        "recommendation",
    }
    known_rec_keys = {"issue", "fix", "priority", "automation_possible"}

    headers = [
        "Manual_Index",
        "item_type",
        "rule",
        "category",
        "status",
        "wcag",
        "wcag_name",
        "wcag_level",
        "impact",
        "recommendation.issue",
        "recommendation.priority",
        "recommendation.automation_possible",
        "recommendation.fix",
        "value",
        "extra_manual_fields_json",
    ]

    rows: List[Dict[str, Any]] = []
    for idx, item in enumerate(mrr):
        if isinstance(item, dict):
            rec = item.get("recommendation", {}) if isinstance(item.get("recommendation", {}), dict) else {}
            fix_val = rec.get("fix", [])
            if isinstance(fix_val, list):
                fix_text = "\n".join(stringify_scalar(x) for x in fix_val)
            else:
                fix_text = stringify_scalar(fix_val)

            extra_item = {k: v for k, v in item.items() if k not in known_keys}
            extra_rec = {k: v for k, v in rec.items() if k not in known_rec_keys}
            extra_payload = {}
            if extra_item:
                extra_payload["extra_manual_item_fields"] = extra_item
            if extra_rec:
                extra_payload["extra_recommendation_fields"] = extra_rec

            rows.append(
                {
                    "Manual_Index": idx,
                    "item_type": "object",
                    "rule": item.get("rule", ""),
                    "category": item.get("category", ""),
                    "status": item.get("status", ""),
                    "wcag": item.get("wcag", ""),
                    "wcag_name": item.get("wcag_name", ""),
                    "wcag_level": item.get("wcag_level", ""),
                    "impact": item.get("impact", ""),
                    "recommendation.issue": rec.get("issue", ""),
                    "recommendation.priority": rec.get("priority", ""),
                    "recommendation.automation_possible": rec.get("automation_possible", ""),
                    "recommendation.fix": fix_text,
                    "value": "",
                    "extra_manual_fields_json": json.dumps(extra_payload, ensure_ascii=False),
                }
            )
        else:
            rows.append(
                {
                    "Manual_Index": idx,
                    "item_type": "string",
                    "rule": "",
                    "category": "",
                    "status": "",
                    "wcag": "",
                    "wcag_name": "",
                    "wcag_level": "",
                    "impact": "",
                    "recommendation.issue": "",
                    "recommendation.priority": "",
                    "recommendation.automation_possible": "",
                    "recommendation.fix": "",
                    "value": stringify_scalar(item),
                    "extra_manual_fields_json": "",
                }
            )

    return headers, rows


def build_overview_rows(report: Dict[str, Any]) -> Tuple[List[str], List[Dict[str, Any]]]:
    headers = ["Metric", "Value"]
    overview_rows: List[Dict[str, Any]] = []

    issues = report.get("issues", [])
    if isinstance(issues, list):
        overview_rows.append({"Metric": "Issues_Count", "Value": str(len(issues))})
    else:
        overview_rows.append({"Metric": "Issues_Count", "Value": "0"})

    mrr = report.get("manual_review_required", [])
    if isinstance(mrr, list):
        overview_rows.append({"Metric": "Manual_Review_Required_Count", "Value": str(len(mrr))})
    else:
        overview_rows.append({"Metric": "Manual_Review_Required_Count", "Value": "0"})

    summary = report.get("summary", {})
    if isinstance(summary, dict):
        for k in ["passed", "failed", "manual_check"]:
            if k in summary:
                overview_rows.append({"Metric": f"Summary_{k}", "Value": stringify_scalar(summary.get(k))})

    # Include other top-level scalar fields.
    for k, v in report.items():
        if k in {"issues", "manual_review_required", "summary"}:
            continue
        if isinstance(v, (dict, list)):
            continue
        overview_rows.append({"Metric": str(k), "Value": stringify_scalar(v)})

    return headers, overview_rows


def auto_fit_columns(ws, max_width: int = 60) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
            cell = row[0]
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), max_width)


def write_table(ws, headers: List[str], rows: List[Dict[str, Any]]) -> None:
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    auto_fit_columns(ws)


def convert_one_report(json_path: Path, output_folder: Path) -> Path:
    with json_path.open("r", encoding="utf-8") as f:
        report = json.load(f)

    doc_name = report.get("document_name")
    # If "document_name" is like "CCJP0619.json", use that without extension.
    if isinstance(doc_name, str) and doc_name.lower().endswith(".json"):
        xlsx_stem = Path(doc_name).stem
    else:
        xlsx_stem = json_path.stem.replace("_Final_report", "")

    output_folder.mkdir(parents=True, exist_ok=True)
    out_path = output_folder / f"{xlsx_stem}.xlsx"

    overview_headers, overview_rows = build_overview_rows(report)
    issues_headers, issues_rows = extract_issues_table(report)
    manual_headers, manual_rows = extract_manual_review_table(report)
    all_fields_rows = build_all_fields_rows(report)

    raw_json_pretty = json.dumps(report, ensure_ascii=False, indent=2)

    wb = Workbook()
    # Workbook starts with a default sheet; we'll remove it.
    default_ws = wb.active
    wb.remove(default_ws)

    ws_overview = wb.create_sheet(safe_sheet_name("Report_Overview"))
    write_table(ws_overview, overview_headers, overview_rows)

    ws_issues = wb.create_sheet(safe_sheet_name("Issues"))
    write_table(ws_issues, issues_headers, issues_rows)

    ws_manual = wb.create_sheet(safe_sheet_name("Manual_Review_Required"))
    write_table(ws_manual, manual_headers, manual_rows)

    ws_all_fields = wb.create_sheet(safe_sheet_name("All_Fields"))
    write_table(ws_all_fields, ["json_path", "value"], all_fields_rows)

    ws_raw = wb.create_sheet(safe_sheet_name("Raw_JSON"))
    ws_raw.append(["raw_json"])
    ws_raw.append([raw_json_pretty])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    ws_raw["A1"].fill = header_fill
    ws_raw["A1"].font = header_font
    ws_raw["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_raw.freeze_panes = "A2"
    auto_fit_columns(ws_raw)

    wb.save(out_path)
    return out_path


def collect_json_reports(input_folder: Path, pattern: str) -> List[Path]:
    if not input_folder.exists() or not input_folder.is_dir():
        raise FileNotFoundError(f"Input folder not found: {input_folder}")

    # Pattern like "*_Final_report.json"
    reports = sorted([p for p in input_folder.glob(pattern) if p.is_file()])
    return reports


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert Adobe 'final' JSON reports to one Excel workbook per JSON."
    )
    parser.add_argument(
        "-i",
        "--input",
        default="Final_reports",
        help="Folder containing *_Final_report.json files (default: Final_reports)",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="Final_report_excels",
        help="Folder to write Excel files (default: Final_report_excels)",
    )
    parser.add_argument(
        "--pattern",
        default="*_Final_report.json",
        help="Glob pattern for JSON reports (default: *_Final_report.json)",
    )

    args = parser.parse_args()

    input_folder = Path(args.input)
    output_folder = Path(args.output)

    reports = collect_json_reports(input_folder, args.pattern)
    if not reports:
        raise FileNotFoundError(
            f"No JSON reports matched pattern '{args.pattern}' in folder: {input_folder}"
        )

    converted_paths: List[str] = []
    for json_path in reports:
        out_path = convert_one_report(json_path, output_folder)
        converted_paths.append(str(out_path))

    print(f"Converted {len(converted_paths)} report(s) -> {output_folder}")


if __name__ == "__main__":
    main()

