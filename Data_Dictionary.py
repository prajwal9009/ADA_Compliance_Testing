import os
import re
import json
import time
import logging
import requests
from datetime import datetime
from pypdf import PdfReader
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
#  CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
# Azure Document Intelligence
AZURE_DOC_ENDPOINT = os.environ.get("AZURE_DOC_INTELLIGENCE_ENDPOINT", "")
AZURE_DOC_KEY      = os.environ.get("AZURE_DOC_INTELLIGENCE_KEY", "")

# Azure OpenAI
AZURE_OPENAI_ENDPOINT    = os.environ.get("AZURE_OPENAI_ENDPOINT", "")
AZURE_OPENAI_API_KEY     = os.environ.get("AZURE_OPENAI_API_KEY", "")
AZURE_OPENAI_API_VERSION = os.environ.get("AZURE_OPENAI_API_VERSION", "2024-02-01")
AZURE_OPENAI_DEPLOYMENT  = os.environ.get("AZURE_OPENAI_DEPLOYMENT", "gpt-4o")

_SCRIPT_DIR       = os.path.dirname(os.path.abspath(__file__))
INPUT_REPORTS_DIR = os.path.join(_SCRIPT_DIR, "Input_reports")
OUTPUT_DIR        = os.path.join(_SCRIPT_DIR, "data Dictionary")

log = logging.getLogger("data_dictionary")


def configure_run_logging(output_dir: str) -> str:
    """
    Write the same messages shown on the console to a timestamped log file
    under <output_dir>/logs/.
    """
    logs_dir = os.path.join(output_dir, "logs")
    os.makedirs(logs_dir, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = os.path.join(logs_dir, f"Data_Dictionary_{stamp}.log")

    log.handlers.clear()
    log.setLevel(logging.INFO)
    log.propagate = False
    fmt = logging.Formatter("%(asctime)s %(levelname)s %(message)s",
                            datefmt="%Y-%m-%d %H:%M:%S")
    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setFormatter(fmt)
    sh = logging.StreamHandler()
    sh.setFormatter(fmt)
    log.addHandler(fh)
    log.addHandler(sh)

    log.info("Log file: %s", log_path)
    return log_path

AZURE_API_VERSION = "2023-07-31"
POLL_INTERVAL_SEC = 2
MAX_POLL_ATTEMPTS = 30

_FIELD_TYPE_MAP = {
    "/Tx":  "Text",
    "/Btn": "Checkbox",
    "/Ch":  "Dropdown",
    "/Sig": "Signature",
}

_UNDEFINED_VALUES = {"", "undefined", "none", "null", "n/a", "na", "-"}
_GENERIC_TOOLTIP_VALUES = {
    "label", "tooltip", "field", "text", "textbox", "input", "value",
    "name", "title", "description", "required"
}


def _acro_tooltip_from_annot(annot) -> str:
    """PDF alternate name / mapping name (common tooltip sources in AcroForm)."""
    for key in ("/TU", "/TM"):
        raw = annot.get(key)
        if raw:
            t = _clean_str(raw)
            if t and not _is_undefined(t):
                return t
    return ""


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 1 — Extract ONLY real AcroForm fields via pypdf (strict gate)
# ─────────────────────────────────────────────────────────────────────────────

def extract_acroform_fields(pdf_path: str) -> list:
    reader     = PdfReader(pdf_path)
    fields     = []
    seen_names = set()

    for page_num, page in enumerate(reader.pages, start=1):
        annots = page.get("/Annots")
        if not annots:
            continue
        for ref in annots:
            annot  = ref.get_object()
            raw_ft = annot.get("/FT")
            if not raw_ft:
                continue
            ft_str = str(raw_ft)
            if ft_str not in _FIELD_TYPE_MAP:
                continue
            raw_name   = annot.get("/T")
            if not raw_name:
                continue
            field_name = _clean_str(raw_name)
            if not field_name or field_name in seen_names:
                continue
            seen_names.add(field_name)

            field_type = _FIELD_TYPE_MAP[ft_str]
            if ft_str == "/Btn":
                ff = annot.get("/Ff")
                if ff and (int(str(ff)) & (1 << 15)):
                    field_type = "Radio Button"

            raw_val = annot.get("/V")
            value   = _clean_str(raw_val) if raw_val else ""

            rect_obj = annot.get("/Rect")
            rect     = [float(v) for v in rect_obj] if rect_obj else []

            fields.append({
                "page":        page_num,
                "field_name":  field_name,
                "field_type":  field_type,
                "field_value": value,
                "pdf_tooltip": _acro_tooltip_from_annot(annot),
                "rect":        rect,
            })

    log.info("  Found %s AcroForm field(s) in the PDF.", len(fields))
    return fields


def _clean_str(val) -> str:
    if val is None:
        return ""
    try:
        if hasattr(val, "to_unicode"):
            return val.to_unicode().strip("()")
        return str(val).strip("()/")
    except Exception:
        return str(val).strip("()/")


def _is_undefined(value: str) -> bool:
    return str(value or "").strip().lower() in _UNDEFINED_VALUES


def _is_meaningful_tooltip(value: str) -> bool:
    txt = str(value or "").strip()
    low = txt.lower()
    if not txt or low in _UNDEFINED_VALUES or low in _GENERIC_TOOLTIP_VALUES:
        return False
    # Single-token placeholders like "label1" or "field_2" are usually not useful.
    if len(txt.split()) == 1 and re.fullmatch(r"[a-zA-Z_]+\d*", txt):
        return False
    return True


def _normalize_identifier(text: str, fallback: str) -> str:
    parts = re.findall(r"[A-Za-z0-9]+", str(text or ""))
    if not parts:
        return fallback
    normalized = "_".join(parts)
    return normalized[:80]


def _fallback_description(field_id: str, field_type: str, pdf_tooltip: str, recommended_label: str = "") -> str:
    """Guaranteed non-empty description from field id/type/pdf tooltip."""
    fid = str(field_id or "").strip() or "this field"
    ftype = str(field_type or "").strip() or "input"
    tip = str(pdf_tooltip or "").strip()
    label = str(recommended_label or "").strip()
    if _is_meaningful_tooltip(tip):
        return f"This {ftype.lower()} field captures {tip.lower()}."
    if label:
        return f"This {ftype.lower()} field captures {label.lower()}."
    return f"This {ftype.lower()} field captures information for {fid}."


def _fallback_recommended_label(field_id: str, pdf_tooltip: str) -> str:
    if _is_meaningful_tooltip(pdf_tooltip):
        return str(pdf_tooltip).strip()
    txt = str(field_id or "").strip()
    if not txt:
        return "Form Field"
    txt = re.sub(r"[_\-]+", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt[:80]


def _fallback_recommended_tooltip(field_id: str, field_type: str, pdf_tooltip: str) -> str:
    if _is_meaningful_tooltip(pdf_tooltip):
        return f'Enter the value for "{str(pdf_tooltip).strip()}".'
    fid = str(field_id or "").strip() or "this field"
    ftype = str(field_type or "").strip() or "field"
    return f"Enter the required value for {fid} ({ftype})."


def _collect_pypdf_label_candidates(pdf_path: str) -> dict:
    """Build page-wise candidate labels from PdfReader extracted text."""
    reader = PdfReader(pdf_path)
    by_page = {}
    for idx, page in enumerate(reader.pages, start=1):
        txt = page.extract_text() or ""
        lines = []
        for line in txt.splitlines():
            cleaned = line.strip()
            if cleaned and len(cleaned) <= 64:
                lines.append(cleaned)
        by_page[idx] = lines
    return by_page


def _ensure_unique_field_names(acro_fields: list) -> list:
    """Guarantee unique field_name values while preserving original text."""
    seen = {}
    for i, field in enumerate(acro_fields, start=1):
        base = str(field.get("field_name", "") or "").strip()
        if not base:
            base = f"Field_{i}"
        count = seen.get(base, 0)
        if count == 0:
            unique_name = base
        else:
            unique_name = f"{base}_{count + 1}"
        seen[base] = count + 1
        field["field_name"] = unique_name
    return acro_fields


def resolve_missing_field_identity(acro_fields: list, pdf_path: str) -> list:
    """
    Use PdfReader/AcroForm as the source of truth for field_name.
    If a field_name is missing/undefined, recover from PdfReader page text.
    Also guarantees there are no duplicate field names.
    """
    pdf_candidates = _collect_pypdf_label_candidates(pdf_path)
    page_cursor = {}

    for i, field in enumerate(acro_fields, start=1):
        page = field.get("page", 1)
        needs_name = _is_undefined(field.get("field_name", ""))
        if not needs_name:
            continue

        candidate = ""
        p_cursor = page_cursor.get(page, 0)
        page_pdf = pdf_candidates.get(page, [])
        if p_cursor < len(page_pdf):
            candidate = page_pdf[p_cursor]
            page_cursor[page] = p_cursor + 1

        if _is_undefined(candidate):
            candidate = f"Field {i}"

        field["field_name"] = str(candidate).strip() or f"Field_{i}"

    return _ensure_unique_field_names(acro_fields)


def get_page_heights(pdf_path: str) -> dict:
    reader = PdfReader(pdf_path)
    return {i + 1: float(page.mediabox.height) for i, page in enumerate(reader.pages)}


# Revision stamps vary by form, e.g.:
#   CCP-0301A (12/01/24)     — hyphen + optional letter before date
#   CCP 0006 (12/01/24)      — space instead of hyphen
#   (12/01/24) CCP 0006      — date first (some single-page forms)
_REVISION_STD = re.compile(
    r"CCP[\s-](?P<n>\d{4})(?P<s>[A-Z]?)\s*\(\s*(?P<d>\d{1,2}/\d{1,2}/\d{2,4})\s*\)",
    re.IGNORECASE,
)
_REVISION_DATE_FIRST = re.compile(
    r"\(\s*(?P<d>\d{1,2}/\d{1,2}/\d{2,4})\s*\)\s*CCP[\s-]?(?P<n>\d{4})(?P<s>[A-Z]?)",
    re.IGNORECASE,
)


def _format_ccp_revision(n: str, s: str, d: str) -> str:
    s = (s or "").strip().upper()
    return f"CCP-{n}{s} ({d})"


def _revision_candidates(text: str) -> list[tuple[str, int]]:
    """Return (normalized stamp, start index) for each match."""
    out = []
    for rx in (_REVISION_STD, _REVISION_DATE_FIRST):
        for m in rx.finditer(text):
            n = m.group("n")
            s = m.group("s") or ""
            d = m.group("d")
            out.append((_format_ccp_revision(n, s, d), m.start()))
    # Same span might not duplicate; dedupe by start index keeping first
    seen = set()
    deduped = []
    for item in sorted(out, key=lambda x: x[1]):
        if item[1] in seen:
            continue
        seen.add(item[1])
        deduped.append(item)
    return deduped


def extract_revision_number_and_date(
    pdf_path: str,
    page_num: int,
    page_texts: dict | None = None,
) -> str:
    """
    Extract revision id and date from the top area of a page (e.g. corner stamp).
    Uses Azure page text when available, otherwise pypdf text for that page.
    """
    text = ""
    if page_texts and page_texts.get(page_num):
        text = page_texts[page_num]
    else:
        try:
            reader = PdfReader(pdf_path)
            if 1 <= page_num <= len(reader.pages):
                text = reader.pages[page_num - 1].extract_text() or ""
        except Exception:
            text = ""
    if not text:
        return ""
    candidates = _revision_candidates(text)
    if not candidates:
        return ""
    if len(candidates) == 1:
        return candidates[0][0]
    header_len = max(400, len(text) // 3)
    for norm, pos in candidates:
        if pos < header_len:
            return norm
    return candidates[0][0]


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 2 — Azure Document Intelligence: page layout + full text
# ─────────────────────────────────────────────────────────────────────────────

def analyze_pdf_with_azure(pdf_path: str) -> dict:
    log.info("  -> Azure Document Intelligence ...")
    url = (
        f"{AZURE_DOC_ENDPOINT.rstrip('/')}/formrecognizer/documentModels/"
        f"prebuilt-document:analyze?api-version={AZURE_API_VERSION}"
    )
    with open(pdf_path, "rb") as fh:
        resp = requests.post(
            url,
            headers={"Ocp-Apim-Subscription-Key": AZURE_DOC_KEY, "Content-Type": "application/pdf"},
            data=fh, timeout=60
        )
    resp.raise_for_status()
    op_url = resp.headers.get("Operation-Location")
    if not op_url:
        raise RuntimeError("Azure did not return an Operation-Location header.")

    for attempt in range(MAX_POLL_ATTEMPTS):
        r = requests.get(op_url, headers={"Ocp-Apim-Subscription-Key": AZURE_DOC_KEY}, timeout=30)
        r.raise_for_status()
        data = r.json()
        if data.get("status") == "succeeded":
            return data
        if data.get("status") == "failed":
            raise RuntimeError(f"Azure analysis failed: {data}")
        log.info("  [%s/%s] %s ...", attempt + 1, MAX_POLL_ATTEMPTS, data.get("status"))
        time.sleep(POLL_INTERVAL_SEC)
    raise TimeoutError("Azure analysis timed out.")


def extract_page_text(azure_result: dict) -> dict:
    """Return {page_number: full_text_string} from Azure lines."""
    page_texts = {}
    for page in azure_result.get("analyzeResult", {}).get("pages", []):
        pnum = page.get("pageNumber", 1)
        lines = [ln.get("content", "") for ln in page.get("lines", []) if ln.get("content", "").strip()]
        page_texts[pnum] = "\n".join(lines)
    return page_texts


def extract_headings(azure_result: dict) -> list:
    """Return list of (page_number, heading_text) tuples from Azure paragraphs."""
    headings = []
    for para in azure_result.get("analyzeResult", {}).get("paragraphs", []):
        if para.get("role") in ("title", "sectionHeading", "heading"):
            text    = para.get("content", "").strip()
            regions = para.get("boundingRegions") or []
            pnum    = regions[0].get("pageNumber", 1) if regions else 1
            if text:
                headings.append((pnum, text))
    return headings


def _poly(obj: dict):
    regions = obj.get("boundingRegions") or []
    if not regions:
        return None
    poly = regions[0].get("polygon") or []
    if len(poly) < 8:
        return None
    xs = poly[0::2]
    ys = poly[1::2]
    return min(xs), min(ys), max(xs), max(ys)


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 3 — Azure OpenAI: generate Label, Description, Tooltip, Section
# ─────────────────────────────────────────────────────────────────────────────

def _call_openai(system_prompt: str, user_prompt: str) -> str:
    """Send a chat completion request to Azure OpenAI and return the text response."""
    url = (
        f"{AZURE_OPENAI_ENDPOINT.rstrip('/')}/openai/deployments/"
        f"{AZURE_OPENAI_DEPLOYMENT}/chat/completions"
        f"?api-version={AZURE_OPENAI_API_VERSION}"
    )
    payload = {
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user",   "content": user_prompt},
        ],
        "temperature": 0,
        "max_tokens":  4000,
    }
    resp = requests.post(
        url,
        headers={"api-key": AZURE_OPENAI_API_KEY, "Content-Type": "application/json"},
        json=payload, timeout=120
    )
    resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"].strip()


def _parse_json_response(raw: str) -> list:
    """Strip markdown fences and parse JSON array from OpenAI response."""
    cleaned = re.sub(r"^```(?:json)?\s*", "", raw.strip(), flags=re.MULTILINE)
    cleaned = re.sub(r"```\s*$",          "", cleaned.strip(), flags=re.MULTILINE)
    cleaned = cleaned.strip()
    try:
        parsed = json.loads(cleaned)
        return parsed if isinstance(parsed, list) else []
    except json.JSONDecodeError:
        # Recover the first bracketed JSON block when the model emits
        # extra prose or partially malformed content.
        match = re.search(r"\[[\s\S]*\]", cleaned)
        if match:
            try:
                parsed = json.loads(match.group(0))
                return parsed if isinstance(parsed, list) else []
            except json.JSONDecodeError:
                pass
        log.warning("  OpenAI field enrichment JSON was invalid; using fallback values.")
        return []


def enrich_fields_with_openai(acro_fields: list, page_texts: dict, headings: list) -> list:
    """
    Single batched OpenAI call.  For every AcroForm field, GPT-4 returns:
        label       – short human-readable label (2–5 words)
        description – one sentence explaining what the field represents
        tooltip     – one instruction sentence for the person filling the form
        section     – which section of the document this field belongs to

    Page text and heading list are passed as context so the model can reason
    about the document structure.
    """
    log.info("  -> Azure OpenAI: generating recommendations + descriptions ...")

    # Build compact field list for the prompt
    field_list = [
        {
            "field_name": f["field_name"],
            "type": f["field_type"],
            "page": f["page"],
            "existing_pdf_tooltip": f.get("pdf_tooltip", ""),
        }
        for f in acro_fields
    ]

    # Include all page text (truncated to stay within token limits)
    full_text = "\n\n--- Page break ---\n\n".join(
        f"[Page {p}]\n{t}" for p, t in sorted(page_texts.items())
    )[:6000]

    heading_list = "\n".join(f"  Page {p}: {h}" for p, h in headings) or "  (none detected)"

    system_prompt = (
        "You are a legal document analyst specializing in court forms. "
        "You receive a list of PDF form fields and the document text, and you "
        "return structured metadata for each field. Always respond with a valid "
        "JSON array and nothing else."
    )

    # Strict ordered list of valid section names
    section_options = "\n".join(f"  - {h}" for _, h in headings) or "  - Form Fields"

    user_prompt = f"""Document text:
{full_text}

Valid section names (use ONLY these — do not invent new ones):
{section_options}

Form fields to analyse:
{json.dumps(field_list, indent=2)}

For EVERY field in the list above return a JSON array where each object has exactly these keys:
  "field_name"               : copy the exact field_name from the input — do not change it
  "recommended_title"        : short business-friendly title (2-6 words), title case
  "recommended_field_name"   : suggested standardized field id (snake_case), based on document intent
  "recommended_label"        : short human-readable label, 2-5 words, title case
  "recommended_tooltip"      : one instruction sentence telling the user how to fill this field
  "description"              : one sentence based ONLY on field_name, type, and existing_pdf_tooltip.
                               Do not leave it blank.
  "section"                  : choose the most appropriate name from the "Valid section names" list above.
                               Read the document text to understand which section each field belongs to.
                               Every field must be assigned one of the listed section names exactly as written.

Return ONLY the JSON array. No explanation, no markdown fences."""

    raw      = _call_openai(system_prompt, user_prompt)
    enriched = _parse_json_response(raw)

    # Index the OpenAI result by field_name for merging
    openai_map = {e.get("field_name", ""): e for e in enriched}

    result = []
    for f in acro_fields:
        oai = openai_map.get(f["field_name"], {})
        recommended_label = oai.get("recommended_label", "") or oai.get("label", "")
        if _is_undefined(recommended_label):
            recommended_label = _fallback_recommended_label(
                f.get("field_name", ""),
                f.get("pdf_tooltip", ""),
            )
        recommended_tooltip = oai.get("recommended_tooltip", "") or oai.get("tooltip", "")
        if _is_undefined(recommended_tooltip):
            recommended_tooltip = _fallback_recommended_tooltip(
                f.get("field_name", ""),
                f.get("field_type", ""),
                f.get("pdf_tooltip", ""),
            )
        recommended_title = oai.get("recommended_title", "")
        if _is_undefined(recommended_title):
            recommended_title = recommended_label
        recommended_field_name = oai.get("recommended_field_name", "")
        if _is_undefined(recommended_field_name):
            recommended_field_name = _normalize_identifier(f["field_name"], fallback="field_name")
        description = oai.get("description", "")
        if _is_undefined(description):
            description = _fallback_description(
                f.get("field_name", ""),
                f.get("field_type", ""),
                f.get("pdf_tooltip", ""),
                recommended_label,
            )
        result.append({
            "page":        f["page"],
            "section":     oai.get("section", "Form Fields"),
            "field_name":  f["field_name"],
            "label":       f["field_name"],
            "type":        f["field_type"],
            "description": description,
            "existing_pdf_tooltip": f.get("pdf_tooltip", ""),
            "recommended_title": recommended_title,
            "recommended_field_name": recommended_field_name,
            "recommended_label": recommended_label,
            "recommended_tooltip": recommended_tooltip,
        })

    log.info("  OpenAI enrichment complete.")
    return result


def assign_sr_nos(fields: list) -> list:
    counters = {}
    for field in fields:
        s = field["section"]
        counters[s] = counters.get(s, 0) + 1
        field["sr_no"] = counters[s]
    return fields


# ─────────────────────────────────────────────────────────────────────────────
#  METADATA EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def extract_pdf_metadata(pdf_path: str) -> dict:
    """Read embedded PDF properties. Author is only set if explicitly present."""
    meta = {
        "file_name":      os.path.splitext(os.path.basename(pdf_path))[0],
        "document_title": "",
        "author":         "",
        "subject":        "",
        "keywords":       "",
        "language":       "",
    }
    try:
        reader = PdfReader(pdf_path)
        info   = reader.metadata or {}

        def pick(*keys) -> str:
            for k in keys:
                for v in (k, k.lower(), k.upper()):
                    val = info.get(v)
                    if val:
                        return str(val).strip()
            return ""

        meta["document_title"] = pick("/Title",    "Title")
        meta["author"]         = pick("/Author",   "Author")
        meta["subject"]        = pick("/Subject",  "Subject")
        meta["keywords"]       = pick("/Keywords", "Keywords")

        # /Lang in the document catalog (e.g. "en-US")
        try:
            root = reader.trailer.get("/Root") or {}
            lang = root.get("/Lang")
            if lang:
                meta["language"] = str(lang).strip("/").strip()
        except Exception:
            pass

        # XMP dc:language as fallback
        if not meta["language"]:
            xmp = reader.xmp_metadata
            if xmp:
                try:
                    lang = xmp.dc_language
                    if lang:
                        meta["language"] = ", ".join(lang) if isinstance(lang, list) else str(lang)
                except Exception:
                    pass

    except Exception as exc:
        log.warning("  could not read PDF metadata: %s", exc)
    return meta


def enrich_metadata_with_openai(meta: dict, page_texts: dict, headings: list) -> dict:
    """
    Use OpenAI to generate accurate Subject, Keywords, and detect Language
    from the actual document content.
    Author is never inferred — only kept if present in the PDF Info dict.
    """
    log.info("  -> Azure OpenAI: generating metadata ...")

    system_prompt = (
        "You are a legal document metadata specialist. "
        "Given document content, return a JSON object with accurate metadata. "
        "Respond with valid JSON only."
    )

    # Numbered list of every heading so the model can reference them by index
    numbered_headings = "\n".join(f"  [{i+1}] {h}" for i, (_, h) in enumerate(headings))
    if not numbered_headings:
        numbered_headings = "  (no headings detected)"

    # Use full extracted content (all pages), chunked to keep prompt size manageable.
    all_pages_text = "\n\n".join(
        f"[Page {p}]\n{t}" for p, t in sorted(page_texts.items()) if t and t.strip()
    ).strip()
    if not all_pages_text:
        all_pages_text = "(no extracted page text)"

    chunk_size = 12000
    content_chunks = [
        all_pages_text[i:i + chunk_size] for i in range(0, len(all_pages_text), chunk_size)
    ] or [all_pages_text]
    content_chunks = content_chunks[:3]

    merged = {}
    for idx, chunk in enumerate(content_chunks, start=1):
        user_prompt = f"""All headings found in this document (in order):
{numbered_headings}

Document text chunk {idx}/{len(content_chunks)}:
{chunk}

Currently known metadata from PDF properties:
  document_title: {meta.get("document_title", "(none yet)")}
  subject: {meta.get("subject", "(none yet)")}
  keywords: {meta.get("keywords", "(none yet)")}
  language: {meta.get("language", "(none yet)")}

Return a JSON object with exactly these keys:
  "document_title" : meaningful intent-based title (6-12 words) that clearly states what this document is for.
                     Use content understanding, not just copied heading text. No generic titles.
  "subject"        : concise subject phrase representing the document's overall purpose
  "keywords"       : comma-separated list of 8-12 significant legal/domain terms from this document
  "language"       : ISO 639-1 language code (e.g. "en", "es", "fr")

Return ONLY the JSON object. No explanation, no markdown fences."""

        raw = _call_openai(system_prompt, user_prompt)
        cleaned = re.sub(r"^```(?:json)?\s*", "", raw.strip(), flags=re.MULTILINE)
        cleaned = re.sub(r"```\s*$", "", cleaned.strip(), flags=re.MULTILINE)
        cleaned = cleaned.strip()
        try:
            oai = json.loads(cleaned)
        except json.JSONDecodeError:
            match = re.search(r"\{[\s\S]*\}", cleaned)
            if match:
                try:
                    oai = json.loads(match.group(0))
                except json.JSONDecodeError:
                    oai = {}
            else:
                oai = {}
            if not oai:
                log.warning("  OpenAI metadata JSON was invalid for chunk %s.", idx)

        if isinstance(oai, dict):
            for key in ("document_title", "subject", "keywords", "language"):
                val = str(oai.get(key, "")).strip()
                if val and key not in merged:
                    merged[key] = val

    # Prefer intent-based OpenAI title when available; keep extracted title only as fallback.
    if merged.get("document_title", ""):
        meta["document_title"] = merged.get("document_title", "")
    if not meta["subject"]:
        meta["subject"] = merged.get("subject", "")
    if not meta["keywords"]:
        meta["keywords"] = merged.get("keywords", "")
    if not meta["language"]:
        meta["language"] = merged.get("language", "")

    return meta


def enrich_metadata_from_azure_only(meta: dict, azure_result: dict) -> dict:
    """
    Fallback when OpenAI is not available.
    Fills fields from Azure layout analysis and a simple language heuristic.
    """
    analyze  = azure_result.get("analyzeResult", {})
    headings = []
    for para in analyze.get("paragraphs", []):
        if para.get("role") in ("title", "sectionHeading"):
            regions = para.get("boundingRegions") or []
            if regions and regions[0].get("pageNumber", 1) == 1:
                headings.append(para.get("content", "").strip())

    # Document title — first heading on page 1
    if not meta["document_title"] and headings:
        meta["document_title"] = headings[0]

    # Subject — second heading on page 1 (first is the page header)
    if not meta["subject"]:
        if len(headings) >= 2:
            meta["subject"] = headings[1]
        elif headings:
            meta["subject"] = headings[0]

    # Keywords from heading words
    if not meta["keywords"]:
        _STOP = {"a","an","the","and","or","of","to","in","for","on","at","by",
                 "is","are","was","be","with","this","it","from","as","not"}
        freq = {}
        for para in analyze.get("paragraphs", []):
            w = 2 if para.get("role") in ("title","sectionHeading") else 1
            for word in re.findall(r"[A-Za-z]{4,}", para.get("content","")):
                lw = word.lower()
                if lw not in _STOP:
                    freq[lw] = freq.get(lw, 0) + w
        top = sorted(freq, key=lambda x: -freq[x])[:12]
        if top:
            meta["keywords"] = ", ".join(w.title() for w in top)

    # Language — Azure detection array
    if not meta["language"]:
        az_langs = analyze.get("languages", [])
        if az_langs:
            best = max(az_langs, key=lambda l: l.get("confidence", 0))
            meta["language"] = best.get("locale", "").strip()

    # Language — simple English heuristic as last resort
    if not meta["language"]:
        pages = analyze.get("pages", [])
        if pages:
            sample = " ".join(
                ln.get("content", "") for ln in pages[0].get("lines", [])
            ).lower()
            common_en = {"the", "of", "and", "to", "in", "a", "is", "that", "for"}
            words     = set(re.findall(r"\b[a-z]+\b", sample))
            if len(common_en & words) >= 4:
                meta["language"] = "en"

    return meta


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL STYLING
# ─────────────────────────────────────────────────────────────────────────────

_THIN         = Side(style="thin", color="B8CCE4")
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_TITLE_FILL   = PatternFill("solid", fgColor="1F4E79")
_SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
_HEADER_FILL  = PatternFill("solid", fgColor="BDD7EE")
_ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")


def _cell(ws, row, col, value="", fill=None, bold=False,
          size=10, color="000000", align="left", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    if fill:
        c.fill = fill
    c.font      = Font(name="Calibri", bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border    = _BORDER
    return c


# ─────────────────────────────────────────────────────────────────────────────
#  DATA DICTIONARY EXCEL (single consolidated file)
# ─────────────────────────────────────────────────────────────────────────────

DD_COLUMNS = [
    "filename",
    "pageNum",
    "revision number & date",
    "Title",
    "Field ID",
    "Form Field Type",
    "Field Label",
    "Data Type",
    "Required",
    "Description",
    "Existing PDF Tooltip",
    "Recommended Title",
    "Recommended Field Name",
    "Recommended Label",
    "Recommended Tooltip",
]
DD_COL_WIDTHS = [24, 10, 28, 28, 32, 20, 26, 16, 12, 54, 45, 30, 32, 28, 45]


def write_data_dictionary_excel(records: list, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Dictionary"

    for i, width in enumerate(DD_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 18
    for ci, col_name in enumerate(DD_COLUMNS, start=1):
        _cell(ws, 1, ci, col_name, fill=_HEADER_FILL, bold=True, align="center")

    for row_idx, rec in enumerate(records, start=2):
        alt = _ALT_FILL if row_idx % 2 == 0 else None
        ws.row_dimensions[row_idx].height = 16

        row_values = [
            rec.get("filename", ""),
            rec.get("pageNum", ""),
            rec.get("revision number & date", ""),
            rec.get("Title", ""),
            rec.get("Field ID", ""),
            rec.get("Form Field Type", ""),
            rec.get("Field Label", ""),
            rec.get("Data Type", ""),
            rec.get("Required", ""),
            rec.get("Description", ""),
            rec.get("Existing PDF Tooltip", ""),
            rec.get("Recommended Title", ""),
            rec.get("Recommended Field Name", ""),
            rec.get("Recommended Label", ""),
            rec.get("Recommended Tooltip", ""),
        ]
        for ci, value in enumerate(row_values, start=1):
            _cell(ws, row_idx, ci, value, fill=alt, wrap=(ci >= 10))

    ws.freeze_panes = "A2"
    wb.save(output_path)
    log.info("  Saved data dictionary : %s", output_path)


def initialize_data_dictionary_excel(output_path: str):
    """Create the consolidated data dictionary workbook with headers only."""
    write_data_dictionary_excel([], output_path)


def append_data_dictionary_rows(records: list, output_path: str):
    """Append new rows to the existing consolidated data dictionary workbook."""
    if not records:
        return
    if not os.path.exists(output_path):
        initialize_data_dictionary_excel(output_path)

    wb = load_workbook(output_path)
    ws = wb["Data Dictionary"] if "Data Dictionary" in wb.sheetnames else wb.active

    start_row = ws.max_row + 1
    for idx, rec in enumerate(records):
        row_idx = start_row + idx
        alt = _ALT_FILL if row_idx % 2 == 0 else None
        ws.row_dimensions[row_idx].height = 16

        row_values = [
            rec.get("filename", ""),
            rec.get("pageNum", ""),
            rec.get("revision number & date", ""),
            rec.get("Title", ""),
            rec.get("Field ID", ""),
            rec.get("Form Field Type", ""),
            rec.get("Field Label", ""),
            rec.get("Data Type", ""),
            rec.get("Required", ""),
            rec.get("Description", ""),
            rec.get("Existing PDF Tooltip", ""),
            rec.get("Recommended Title", ""),
            rec.get("Recommended Field Name", ""),
            rec.get("Recommended Label", ""),
            rec.get("Recommended Tooltip", ""),
        ]
        for ci, value in enumerate(row_values, start=1):
            _cell(ws, row_idx, ci, value, fill=alt, wrap=(ci >= 10))

    ws.freeze_panes = "A2"
    wb.save(output_path)
    log.info("  Updated data dictionary: %s", output_path)


# ─────────────────────────────────────────────────────────────────────────────
#  METADATA EXCEL
# ─────────────────────────────────────────────────────────────────────────────

_META_PROPS = [
    ("File Name",      "file_name"),
    ("Document Title", "document_title"),
    ("Author",         "author"),
    ("Subject",        "subject"),
    ("Keywords",       "keywords"),
    ("Language",       "language"),
]


def write_metadata_excel(meta: dict, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Metadata"

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 80

    ws.row_dimensions[1].height = 22
    tc = ws.cell(row=1, column=1, value="  Document Metadata")
    tc.fill      = _TITLE_FILL
    tc.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:C1")

    ws.row_dimensions[2].height = 16
    for ci, hdr in enumerate(["Sr No", "Property", "Value"], start=1):
        _cell(ws, 2, ci, hdr, fill=_HEADER_FILL, bold=True, align="center")

    for sr, (label, key) in enumerate(_META_PROPS, start=1):
        alt = _ALT_FILL if sr % 2 == 0 else None
        ws.row_dimensions[sr + 2].height = 15
        _cell(ws, sr + 2, 1, sr,                fill=alt, align="center")
        _cell(ws, sr + 2, 2, label,              fill=alt, bold=True)
        _cell(ws, sr + 2, 3, meta.get(key, ""),  fill=alt, wrap=True)

    wb.save(output_path)
    log.info("  Saved metadata        : %s", output_path)


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def process_pdf(pdf_path: str) -> tuple:
    """
    Pipeline:
      1. pypdf         → strict AcroForm field extraction
      2. Azure Doc Int → page layout, full text, headings
      3. Azure OpenAI  → Label, Description, Tooltip, Section per field
                      → accurate Subject, Keywords, Language for metadata
    """
    # ── Step 1: AcroForm fields ───────────────────────────────────────────────
    acro_fields = extract_acroform_fields(pdf_path)
    if not acro_fields:
        log.info("  No AcroForm fields found in this PDF.")
        return [], {}, {}, []

    # ── Step 2: Azure Document Intelligence ──────────────────────────────────
    azure_result = {}
    page_texts   = {}
    headings     = []

    if AZURE_DOC_ENDPOINT and AZURE_DOC_KEY:
        azure_result = analyze_pdf_with_azure(pdf_path)
        page_texts   = extract_page_text(azure_result)
        headings     = extract_headings(azure_result)
    else:
        log.warning("  Azure Doc Intelligence not configured.")

    # Fill missing field names before OpenAI enrichment:
    # Doc Intelligence candidates first, then PdfReader fallback.
    acro_fields = resolve_missing_field_identity(acro_fields, pdf_path)

    # ── Step 3a: Enrich fields via OpenAI ────────────────────────────────────
    if AZURE_OPENAI_ENDPOINT and AZURE_OPENAI_API_KEY and page_texts:
        enriched = enrich_fields_with_openai(acro_fields, page_texts, headings)
    else:
        if not (AZURE_OPENAI_ENDPOINT and AZURE_OPENAI_API_KEY):
            log.warning("  Azure OpenAI not configured — using field names only.")
        enriched = [
            {
                "page":        f["page"],
                "section":     "Form Fields",
                "field_name":  f["field_name"],
                "label":       f["field_name"],
                "type":        f["field_type"],
                "description": _fallback_description(
                    f.get("field_name", ""),
                    f.get("field_type", ""),
                    f.get("pdf_tooltip", ""),
                    _fallback_recommended_label(
                        f.get("field_name", ""),
                        f.get("pdf_tooltip", ""),
                    ),
                ),
                "existing_pdf_tooltip": f.get("pdf_tooltip", ""),
                "recommended_title": "",
                "recommended_field_name": _normalize_identifier(
                    f.get("field_name", ""), fallback="field_name"
                ),
                "recommended_label": _fallback_recommended_label(
                    f.get("field_name", ""),
                    f.get("pdf_tooltip", ""),
                ),
                "recommended_tooltip": _fallback_recommended_tooltip(
                    f.get("field_name", ""),
                    f.get("field_type", ""),
                    f.get("pdf_tooltip", ""),
                ),
            }
            for f in acro_fields
        ]

    fields = assign_sr_nos(enriched)
    return fields, azure_result, page_texts, headings


def generate_data_dictionary(
    input_reports_dir: str = INPUT_REPORTS_DIR,
    output_dir:        str = OUTPUT_DIR,
):
    os.makedirs(output_dir, exist_ok=True)
    configure_run_logging(output_dir)

    if not os.path.isdir(input_reports_dir):
        os.makedirs(input_reports_dir, exist_ok=True)
        log.info("Created input folder: %s", input_reports_dir)
        log.info("Place your PDF files there and run again.")
        return

    pdf_files = sorted(
        os.path.join(input_reports_dir, f)
        for f in os.listdir(input_reports_dir)
        if f.lower().endswith(".pdf")
    )
    if not pdf_files:
        log.info("No PDF files found in: %s", input_reports_dir)
        return

    log.info("Found %s PDF(s) to process.\n", len(pdf_files))
    dict_file = os.path.join(output_dir, "Generated_Data_Dictionary.xlsx")
    initialize_data_dictionary_excel(dict_file)

    for pdf_path in pdf_files:
        pdf_stem  = os.path.splitext(os.path.basename(pdf_path))[0]
        meta_file = os.path.join(output_dir, f"{pdf_stem}_Metadata.xlsx")

        log.info("Processing: %s.pdf", pdf_stem)
        try:
            fields, azure_result, page_texts, headings = process_pdf(pdf_path)

            # ── Metadata ──────────────────────────────────────────────────────
            meta = extract_pdf_metadata(pdf_path)
            # For PDFs without AcroForm fields, process_pdf() exits before Azure extraction.
            # Run metadata-only extraction so OpenAI can still infer meaningful metadata.
            if (not page_texts) and AZURE_DOC_ENDPOINT and AZURE_DOC_KEY:
                azure_result = analyze_pdf_with_azure(pdf_path)
                page_texts = extract_page_text(azure_result)
                headings = extract_headings(azure_result)

            if AZURE_OPENAI_ENDPOINT and AZURE_OPENAI_API_KEY and page_texts:
                # OpenAI gives accurate Subject, Keywords, Language
                meta = enrich_metadata_with_openai(meta, page_texts, headings)
            elif azure_result:
                # Fallback: derive from layout analysis only
                meta = enrich_metadata_from_azure_only(meta, azure_result)

            resolved_title = meta.get("document_title", "")
            resolved_recommended_title = resolved_title
            revision_by_page = {}
            pdf_dictionary_rows = []
            for f in fields:
                pnum = f.get("page", 1)
                try:
                    pnum = int(pnum)
                except (TypeError, ValueError):
                    pnum = 1
                if pnum not in revision_by_page:
                    revision_by_page[pnum] = extract_revision_number_and_date(
                        pdf_path, pnum, page_texts
                    )
                pdf_dictionary_rows.append({
                    "filename": pdf_stem,
                    "pageNum": f.get("page", ""),
                    "revision number & date": revision_by_page.get(pnum, ""),
                    "Title": resolved_title,
                    "Field ID": f.get("field_name", ""),
                    "Form Field Type": f.get("type", ""),
                    "Field Label": f.get("label", ""),
                    "Data Type": "",
                    "Required": "",
                    "Description": f.get("description", ""),
                    "Existing PDF Tooltip": f.get("existing_pdf_tooltip", ""),
                    "Recommended Title": resolved_recommended_title,
                    "Recommended Field Name": f.get("recommended_field_name", ""),
                    "Recommended Label": f.get("recommended_label", ""),
                    "Recommended Tooltip": f.get("recommended_tooltip", ""),
                })
            append_data_dictionary_rows(pdf_dictionary_rows, dict_file)

            write_metadata_excel(meta, meta_file)
            log.info("Done: %s\n", pdf_stem)

        except Exception as exc:
            log.exception("Error processing %s: %s", pdf_path, exc)
            log.info("")

    log.info("All done.")


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Generate Data Dictionary and Metadata from PDF form fields."
    )
    parser.add_argument("--input",                default=INPUT_REPORTS_DIR)
    parser.add_argument("--output",               default=OUTPUT_DIR)
    parser.add_argument("--doc-endpoint",         default="", help="Azure Doc Intelligence endpoint")
    parser.add_argument("--doc-key",              default="", help="Azure Doc Intelligence API key")
    parser.add_argument("--openai-endpoint",      default="", help="Azure OpenAI endpoint")
    parser.add_argument("--openai-key",           default="", help="Azure OpenAI API key")
    parser.add_argument("--openai-api-version",   default="", help="Azure OpenAI API version")
    parser.add_argument("--openai-deployment",    default="", help="Azure OpenAI deployment name")
    args = parser.parse_args()

    if args.doc_endpoint:      AZURE_DOC_ENDPOINT       = args.doc_endpoint
    if args.doc_key:           AZURE_DOC_KEY            = args.doc_key
    if args.openai_endpoint:   AZURE_OPENAI_ENDPOINT    = args.openai_endpoint
    if args.openai_key:        AZURE_OPENAI_API_KEY     = args.openai_key
    if args.openai_api_version:AZURE_OPENAI_API_VERSION = args.openai_api_version
    if args.openai_deployment: AZURE_OPENAI_DEPLOYMENT  = args.openai_deployment

    generate_data_dictionary(
        input_reports_dir=args.input,
        output_dir=args.output,
    )
