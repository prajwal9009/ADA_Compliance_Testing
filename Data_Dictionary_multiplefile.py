import os
import re
import json
import time
import requests
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
#  CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
# Azure Document Intelligence
AZURE_DOC_ENDPOINT = os.environ.get("AZURE_DOC_INTELLIGENCE_ENDPOINT", "")
AZURE_DOC_KEY      = os.environ.get("AZURE_DOC_INTELLIGENCE_KEY", "")

# Azure OpenAI
AZURE_OPENAI_ENDPOINT    = os.environ.get("AZURE_OPENAI_ENDPOINT", "")
AZURE_OPENAI_API_KEY     = os.environ.get("AZURE_OPENAI_API_KEY", "")
AZURE_OPENAI_API_VERSION = os.environ.get("AZURE_OPENAI_API_VERSION", "")
AZURE_OPENAI_DEPLOYMENT  = os.environ.get("AZURE_OPENAI_DEPLOYMENT", "")

_SCRIPT_DIR       = os.path.dirname(os.path.abspath(__file__))
INPUT_REPORTS_DIR = os.path.join(_SCRIPT_DIR, "Input_reports")
OUTPUT_DIR        = os.path.join(_SCRIPT_DIR, "data Dictionary")

AZURE_API_VERSION = "2023-07-31"
POLL_INTERVAL_SEC = 2
MAX_POLL_ATTEMPTS = 30

_FIELD_TYPE_MAP = {
    "/Tx":  "Text",
    "/Btn": "Checkbox",
    "/Ch":  "Dropdown",
    "/Sig": "Signature",
}

_UNDEFINED_VALUES = {"", "undefined", "none", "null", "n/a", "na", "-"}


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 1 — Extract ONLY real AcroForm fields via pypdf (strict gate)
# ─────────────────────────────────────────────────────────────────────────────

def extract_acroform_fields(pdf_path: str) -> list:
    reader     = PdfReader(pdf_path)
    fields     = []
    seen_names = set()

    for page_num, page in enumerate(reader.pages, start=1):
        annots = page.get("/Annots")
        if not annots:
            continue
        for ref in annots:
            annot  = ref.get_object()
            raw_ft = annot.get("/FT")
            if not raw_ft:
                continue
            ft_str = str(raw_ft)
            if ft_str not in _FIELD_TYPE_MAP:
                continue
            raw_name   = annot.get("/T")
            if not raw_name:
                continue
            field_name = _clean_str(raw_name)
            if not field_name or field_name in seen_names:
                continue
            seen_names.add(field_name)

            field_type = _FIELD_TYPE_MAP[ft_str]
            if ft_str == "/Btn":
                ff = annot.get("/Ff")
                if ff and (int(str(ff)) & (1 << 15)):
                    field_type = "Radio Button"

            raw_val = annot.get("/V")
            value   = _clean_str(raw_val) if raw_val else ""

            rect_obj = annot.get("/Rect")
            rect     = [float(v) for v in rect_obj] if rect_obj else []

            fields.append({
                "page":        page_num,
                "field_name":  field_name,
                "field_type":  field_type,
                "field_value": value,
                "rect":        rect,
            })

    print(f"  Found {len(fields)} AcroForm field(s) in the PDF.")
    return fields


def _clean_str(val) -> str:
    if val is None:
        return ""
    try:
        if hasattr(val, "to_unicode"):
            return val.to_unicode().strip("()")
        return str(val).strip("()/")
    except Exception:
        return str(val).strip("()/")


def _is_undefined(value: str) -> bool:
    return str(value or "").strip().lower() in _UNDEFINED_VALUES


def _normalize_identifier(text: str, fallback: str) -> str:
    parts = re.findall(r"[A-Za-z0-9]+", str(text or ""))
    if not parts:
        return fallback
    normalized = "_".join(parts)
    return normalized[:80]


def _collect_pypdf_label_candidates(pdf_path: str) -> dict:
    """Build page-wise candidate labels from PdfReader extracted text."""
    reader = PdfReader(pdf_path)
    by_page = {}
    for idx, page in enumerate(reader.pages, start=1):
        txt = page.extract_text() or ""
        lines = []
        for line in txt.splitlines():
            cleaned = line.strip()
            if cleaned and len(cleaned) <= 64:
                lines.append(cleaned)
        by_page[idx] = lines
    return by_page


def _ensure_unique_field_names(acro_fields: list) -> list:
    """Guarantee unique field_name values by appending numeric suffixes."""
    seen = {}
    for field in acro_fields:
        base = _normalize_identifier(field.get("field_name", ""), fallback="Field")
        count = seen.get(base, 0)
        if count == 0:
            unique_name = base
        else:
            unique_name = f"{base}_{count + 1}"
        seen[base] = count + 1
        field["field_name"] = unique_name
    return acro_fields


def resolve_missing_field_identity(acro_fields: list, pdf_path: str) -> list:
    """
    Use PdfReader/AcroForm as the source of truth for field_name.
    If a field_name is missing/undefined, recover from PdfReader page text.
    Also guarantees there are no duplicate field names.
    """
    pdf_candidates = _collect_pypdf_label_candidates(pdf_path)
    page_cursor = {}

    for i, field in enumerate(acro_fields, start=1):
        page = field.get("page", 1)
        needs_name = _is_undefined(field.get("field_name", ""))
        if not needs_name:
            continue

        candidate = ""
        p_cursor = page_cursor.get(page, 0)
        page_pdf = pdf_candidates.get(page, [])
        if p_cursor < len(page_pdf):
            candidate = page_pdf[p_cursor]
            page_cursor[page] = p_cursor + 1

        if _is_undefined(candidate):
            candidate = f"Field {i}"

        field["field_name"] = _normalize_identifier(candidate, fallback=f"Field_{i}")

    return _ensure_unique_field_names(acro_fields)


def get_page_heights(pdf_path: str) -> dict:
    reader = PdfReader(pdf_path)
    return {i + 1: float(page.mediabox.height) for i, page in enumerate(reader.pages)}


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 2 — Azure Document Intelligence: page layout + full text
# ─────────────────────────────────────────────────────────────────────────────

def analyze_pdf_with_azure(pdf_path: str) -> dict:
    print(f"  -> Azure Document Intelligence ...")
    url = (
        f"{AZURE_DOC_ENDPOINT.rstrip('/')}/formrecognizer/documentModels/"
        f"prebuilt-document:analyze?api-version={AZURE_API_VERSION}"
    )
    with open(pdf_path, "rb") as fh:
        resp = requests.post(
            url,
            headers={"Ocp-Apim-Subscription-Key": AZURE_DOC_KEY, "Content-Type": "application/pdf"},
            data=fh, timeout=60
        )
    resp.raise_for_status()
    op_url = resp.headers.get("Operation-Location")
    if not op_url:
        raise RuntimeError("Azure did not return an Operation-Location header.")

    for attempt in range(MAX_POLL_ATTEMPTS):
        r = requests.get(op_url, headers={"Ocp-Apim-Subscription-Key": AZURE_DOC_KEY}, timeout=30)
        r.raise_for_status()
        data = r.json()
        if data.get("status") == "succeeded":
            return data
        if data.get("status") == "failed":
            raise RuntimeError(f"Azure analysis failed: {data}")
        print(f"  [{attempt + 1}/{MAX_POLL_ATTEMPTS}] {data.get('status')} ...")
        time.sleep(POLL_INTERVAL_SEC)
    raise TimeoutError("Azure analysis timed out.")


def extract_page_text(azure_result: dict) -> dict:
    """Return {page_number: full_text_string} from Azure lines."""
    page_texts = {}
    for page in azure_result.get("analyzeResult", {}).get("pages", []):
        pnum = page.get("pageNumber", 1)
        lines = [ln.get("content", "") for ln in page.get("lines", []) if ln.get("content", "").strip()]
        page_texts[pnum] = "\n".join(lines)
    return page_texts


def extract_headings(azure_result: dict) -> list:
    """Return list of (page_number, heading_text) tuples from Azure paragraphs."""
    headings = []
    for para in azure_result.get("analyzeResult", {}).get("paragraphs", []):
        if para.get("role") in ("title", "sectionHeading", "heading"):
            text    = para.get("content", "").strip()
            regions = para.get("boundingRegions") or []
            pnum    = regions[0].get("pageNumber", 1) if regions else 1
            if text:
                headings.append((pnum, text))
    return headings


def _poly(obj: dict):
    regions = obj.get("boundingRegions") or []
    if not regions:
        return None
    poly = regions[0].get("polygon") or []
    if len(poly) < 8:
        return None
    xs = poly[0::2]
    ys = poly[1::2]
    return min(xs), min(ys), max(xs), max(ys)


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 3 — Azure OpenAI: generate Label, Description, Tooltip, Section
# ─────────────────────────────────────────────────────────────────────────────

def _call_openai(system_prompt: str, user_prompt: str) -> str:
    """Send a chat completion request to Azure OpenAI and return the text response."""
    url = (
        f"{AZURE_OPENAI_ENDPOINT.rstrip('/')}/openai/deployments/"
        f"{AZURE_OPENAI_DEPLOYMENT}/chat/completions"
        f"?api-version={AZURE_OPENAI_API_VERSION}"
    )
    payload = {
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user",   "content": user_prompt},
        ],
        "temperature": 0,
        "max_tokens":  4000,
    }
    resp = requests.post(
        url,
        headers={"api-key": AZURE_OPENAI_API_KEY, "Content-Type": "application/json"},
        json=payload, timeout=120
    )
    resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"].strip()


def _parse_json_response(raw: str) -> list:
    """Strip markdown fences and parse JSON array from OpenAI response."""
    cleaned = re.sub(r"^```(?:json)?\s*", "", raw.strip(), flags=re.MULTILINE)
    cleaned = re.sub(r"```\s*$",          "", cleaned.strip(), flags=re.MULTILINE)
    return json.loads(cleaned.strip())


def enrich_fields_with_openai(acro_fields: list, page_texts: dict, headings: list) -> list:
    """
    Single batched OpenAI call.  For every AcroForm field, GPT-4 returns:
        label       – short human-readable label (2–5 words)
        description – one sentence explaining what the field represents
        tooltip     – one instruction sentence for the person filling the form
        section     – which section of the document this field belongs to

    Page text and heading list are passed as context so the model can reason
    about the document structure.
    """
    print(f"  -> Azure OpenAI: generating Label / Description / Tooltip / Section ...")

    # Build compact field list for the prompt
    field_list = [
        {
            "field_name": f["field_name"],
            "type": f["field_type"],
            "page": f["page"],
        }
        for f in acro_fields
    ]

    # Include all page text (truncated to stay within token limits)
    full_text = "\n\n--- Page break ---\n\n".join(
        f"[Page {p}]\n{t}" for p, t in sorted(page_texts.items())
    )[:6000]

    heading_list = "\n".join(f"  Page {p}: {h}" for p, h in headings) or "  (none detected)"

    system_prompt = (
        "You are a legal document analyst specializing in court forms. "
        "You receive a list of PDF form fields and the document text, and you "
        "return structured metadata for each field. Always respond with a valid "
        "JSON array and nothing else."
    )

    # Strict ordered list of valid section names
    section_options = "\n".join(f"  - {h}" for _, h in headings) or "  - Form Fields"

    user_prompt = f"""Document text:
{full_text}

Valid section names (use ONLY these — do not invent new ones):
{section_options}

Form fields to analyse:
{json.dumps(field_list, indent=2)}

For EVERY field in the list above return a JSON array where each object has exactly these keys:
  "field_name"  : copy the exact field_name from the input — do not change it
  "label"       : short human-readable label, 2-5 words, title case, inferred from field intent and document context
  "description" : one sentence explaining what information this field captures
  "tooltip"     : one instruction sentence telling the user how to fill this field
  "section"     : choose the most appropriate name from the "Valid section names" list above.
                  Read the document text to understand which section each field belongs to.
                  Every field must be assigned one of the listed section names exactly as written.

Return ONLY the JSON array. No explanation, no markdown fences."""

    raw      = _call_openai(system_prompt, user_prompt)
    enriched = _parse_json_response(raw)

    # Index the OpenAI result by field_name for merging
    openai_map = {e.get("field_name", ""): e for e in enriched}

    result = []
    for f in acro_fields:
        oai = openai_map.get(f["field_name"], {})
        chosen_label = oai.get("label", "")
        if _is_undefined(chosen_label):
            chosen_label = f["field_name"]
        result.append({
            "page":        f["page"],
            "section":     oai.get("section", "Form Fields"),
            "field_name":  f["field_name"],
            "label":       chosen_label,
            "type":        f["field_type"],
            "description": oai.get("description", ""),
            "tooltip":     oai.get("tooltip",     ""),
        })

    print(f"  OpenAI enrichment complete.")
    return result


def assign_sr_nos(fields: list) -> list:
    counters = {}
    for field in fields:
        s = field["section"]
        counters[s] = counters.get(s, 0) + 1
        field["sr_no"] = counters[s]
    return fields


# ─────────────────────────────────────────────────────────────────────────────
#  METADATA EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def extract_pdf_metadata(pdf_path: str) -> dict:
    """Read embedded PDF properties. Author is only set if explicitly present."""
    meta = {
        "file_name":      os.path.splitext(os.path.basename(pdf_path))[0],
        "document_title": "",
        "author":         "",
        "subject":        "",
        "keywords":       "",
        "language":       "",
    }
    try:
        reader = PdfReader(pdf_path)
        info   = reader.metadata or {}

        def pick(*keys) -> str:
            for k in keys:
                for v in (k, k.lower(), k.upper()):
                    val = info.get(v)
                    if val:
                        return str(val).strip()
            return ""

        meta["document_title"] = pick("/Title",    "Title")
        meta["author"]         = pick("/Author",   "Author")
        meta["subject"]        = pick("/Subject",  "Subject")
        meta["keywords"]       = pick("/Keywords", "Keywords")

        # /Lang in the document catalog (e.g. "en-US")
        try:
            root = reader.trailer.get("/Root") or {}
            lang = root.get("/Lang")
            if lang:
                meta["language"] = str(lang).strip("/").strip()
        except Exception:
            pass

        # XMP dc:language as fallback
        if not meta["language"]:
            xmp = reader.xmp_metadata
            if xmp:
                try:
                    lang = xmp.dc_language
                    if lang:
                        meta["language"] = ", ".join(lang) if isinstance(lang, list) else str(lang)
                except Exception:
                    pass

    except Exception as exc:
        print(f"  Warning: could not read PDF metadata: {exc}")
    return meta


def enrich_metadata_with_openai(meta: dict, page_texts: dict, headings: list) -> dict:
    """
    Use OpenAI to generate accurate Subject, Keywords, and detect Language
    from the actual document content.
    Author is never inferred — only kept if present in the PDF Info dict.
    """
    print(f"  -> Azure OpenAI: generating metadata ...")

    first_page_text = page_texts.get(1, "")[:3000]
    heading_list    = "\n".join(f"  {h}" for _, h in headings) or "  (none)"

    system_prompt = (
        "You are a legal document metadata specialist. "
        "Given a document's headings and first-page text, return a JSON object "
        "with accurate metadata. Respond with valid JSON only."
    )

    # Numbered list of every heading so the model can reference them by index
    numbered_headings = "\n".join(f"  [{i+1}] {h}" for i, (_, h) in enumerate(headings))
    if not numbered_headings:
        numbered_headings = "  (no headings detected)"

    user_prompt = f"""All headings found in this document (in order):
{numbered_headings}

First page text:
{first_page_text}

Keywords already extracted: {meta.get("keywords", "(none yet)")}

Return a JSON object with exactly these keys:
  "document_title" : the official title — copy the exact text of the first heading on page 1
  "subject"        : copy the exact text of whichever heading from the list above BEST describes
                     the document's overall legal purpose. Use the keywords and page content to
                     decide — pick the heading that most closely matches the core topic.
                     Do NOT invent or rephrase — return the heading verbatim.
  "keywords"       : comma-separated list of 8-12 significant legal terms from this document
  "language"       : ISO 639-1 language code (e.g. "en", "es", "fr") detected from the text

Return ONLY the JSON object. No explanation, no markdown fences."""

    raw  = _call_openai(system_prompt, user_prompt)
    # Strip fences
    cleaned = re.sub(r"^```(?:json)?\s*", "", raw.strip(), flags=re.MULTILINE)
    cleaned = re.sub(r"```\s*$",          "", cleaned.strip(), flags=re.MULTILINE)
    oai     = json.loads(cleaned.strip())

    if not meta["document_title"]:
        meta["document_title"] = oai.get("document_title", "")
    if not meta["subject"]:
        meta["subject"]  = oai.get("subject",  "")
    if not meta["keywords"]:
        meta["keywords"] = oai.get("keywords", "")
    if not meta["language"]:
        meta["language"] = oai.get("language", "")

    return meta


def enrich_metadata_from_azure_only(meta: dict, azure_result: dict) -> dict:
    """
    Fallback when OpenAI is not available.
    Fills fields from Azure layout analysis and a simple language heuristic.
    """
    analyze  = azure_result.get("analyzeResult", {})
    headings = []
    for para in analyze.get("paragraphs", []):
        if para.get("role") in ("title", "sectionHeading"):
            regions = para.get("boundingRegions") or []
            if regions and regions[0].get("pageNumber", 1) == 1:
                headings.append(para.get("content", "").strip())

    # Document title — first heading on page 1
    if not meta["document_title"] and headings:
        meta["document_title"] = headings[0]

    # Subject — second heading on page 1 (first is the page header)
    if not meta["subject"]:
        if len(headings) >= 2:
            meta["subject"] = headings[1]
        elif headings:
            meta["subject"] = headings[0]

    # Keywords from heading words
    if not meta["keywords"]:
        _STOP = {"a","an","the","and","or","of","to","in","for","on","at","by",
                 "is","are","was","be","with","this","it","from","as","not"}
        freq = {}
        for para in analyze.get("paragraphs", []):
            w = 2 if para.get("role") in ("title","sectionHeading") else 1
            for word in re.findall(r"[A-Za-z]{4,}", para.get("content","")):
                lw = word.lower()
                if lw not in _STOP:
                    freq[lw] = freq.get(lw, 0) + w
        top = sorted(freq, key=lambda x: -freq[x])[:12]
        if top:
            meta["keywords"] = ", ".join(w.title() for w in top)

    # Language — Azure detection array
    if not meta["language"]:
        az_langs = analyze.get("languages", [])
        if az_langs:
            best = max(az_langs, key=lambda l: l.get("confidence", 0))
            meta["language"] = best.get("locale", "").strip()

    # Language — simple English heuristic as last resort
    if not meta["language"]:
        pages = analyze.get("pages", [])
        if pages:
            sample = " ".join(
                ln.get("content", "") for ln in pages[0].get("lines", [])
            ).lower()
            common_en = {"the", "of", "and", "to", "in", "a", "is", "that", "for"}
            words     = set(re.findall(r"\b[a-z]+\b", sample))
            if len(common_en & words) >= 4:
                meta["language"] = "en"

    return meta


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL STYLING
# ─────────────────────────────────────────────────────────────────────────────

_THIN         = Side(style="thin", color="B8CCE4")
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_TITLE_FILL   = PatternFill("solid", fgColor="1F4E79")
_SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
_HEADER_FILL  = PatternFill("solid", fgColor="BDD7EE")
_ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")


def _cell(ws, row, col, value="", fill=None, bold=False,
          size=10, color="000000", align="left", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    if fill:
        c.fill = fill
    c.font      = Font(name="Calibri", bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border    = _BORDER
    return c


# ─────────────────────────────────────────────────────────────────────────────
#  DATA DICTIONARY EXCEL  —  matches reference column structure exactly
#  Sr .No. | Field Name | Label | Type | Description | Tooltip
# ─────────────────────────────────────────────────────────────────────────────

DD_COLUMNS    = ["Sr .No.", "Field Name", "Label", "Type", "Description", "Tooltip"]
DD_COL_WIDTHS = [8, 38, 28, 14, 45, 50]


def write_data_dictionary_excel(fields: list, output_path: str, form_name: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "form data"

    for i, w in enumerate(DD_COL_WIDTHS, start=3):   # starts at column C
        ws.column_dimensions[get_column_letter(i)].width = w

    row_num = 1

    # Title row
    ws.row_dimensions[row_num].height = 22
    tc = ws.cell(row=row_num, column=3,
                 value=f"  Form fields data dictionary  -  {form_name}")
    tc.fill      = _TITLE_FILL
    tc.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row_num, start_column=3,
                   end_row=row_num,   end_column=3 + len(DD_COLUMNS) - 1)
    row_num += 2

    current_section = None
    sr_counter      = 0

    for rec in fields:
        section = rec["section"]

        if section != current_section:
            current_section = section
            sr_counter      = 0

            # Section header
            ws.row_dimensions[row_num].height = 18
            sc = ws.cell(row=row_num, column=3, value=f"  Section :  {section}")
            sc.fill      = _SECTION_FILL
            sc.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
            sc.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=row_num, start_column=3,
                           end_row=row_num,   end_column=3 + len(DD_COLUMNS) - 1)
            row_num += 1

            # Column header row
            ws.row_dimensions[row_num].height = 16
            for ci, col_name in enumerate(DD_COLUMNS, start=3):
                _cell(ws, row_num, ci, col_name,
                      fill=_HEADER_FILL, bold=True, align="center")
            row_num += 1

        sr_counter += 1
        alt = _ALT_FILL if sr_counter % 2 == 0 else None
        ws.row_dimensions[row_num].height = 15

        for ci, val in enumerate([
            sr_counter,
            rec["field_name"],
            rec["label"],
            rec["type"],
            rec.get("description", ""),
            rec.get("tooltip",     ""),
        ], start=3):
            _cell(ws, row_num, ci, val, fill=alt, wrap=(ci >= 7))
        row_num += 1

    ws.freeze_panes = "C6"
    wb.save(output_path)
    print(f"  Saved data dictionary : {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
#  METADATA EXCEL
# ─────────────────────────────────────────────────────────────────────────────

_META_PROPS = [
    ("File Name",      "file_name"),
    ("Document Title", "document_title"),
    ("Author",         "author"),
    ("Subject",        "subject"),
    ("Keywords",       "keywords"),
    ("Language",       "language"),
]


def write_metadata_excel(meta: dict, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Metadata"

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 80

    ws.row_dimensions[1].height = 22
    tc = ws.cell(row=1, column=1, value="  Document Metadata")
    tc.fill      = _TITLE_FILL
    tc.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:C1")

    ws.row_dimensions[2].height = 16
    for ci, hdr in enumerate(["Sr No", "Property", "Value"], start=1):
        _cell(ws, 2, ci, hdr, fill=_HEADER_FILL, bold=True, align="center")

    for sr, (label, key) in enumerate(_META_PROPS, start=1):
        alt = _ALT_FILL if sr % 2 == 0 else None
        ws.row_dimensions[sr + 2].height = 15
        _cell(ws, sr + 2, 1, sr,                fill=alt, align="center")
        _cell(ws, sr + 2, 2, label,              fill=alt, bold=True)
        _cell(ws, sr + 2, 3, meta.get(key, ""),  fill=alt, wrap=True)

    wb.save(output_path)
    print(f"  Saved metadata        : {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def process_pdf(pdf_path: str) -> tuple:
    """
    Pipeline:
      1. pypdf         → strict AcroForm field extraction
      2. Azure Doc Int → page layout, full text, headings
      3. Azure OpenAI  → Label, Description, Tooltip, Section per field
                      → accurate Subject, Keywords, Language for metadata
    """
    # ── Step 1: AcroForm fields ───────────────────────────────────────────────
    acro_fields = extract_acroform_fields(pdf_path)
    if not acro_fields:
        print("  No AcroForm fields found in this PDF.")
        return [], {}

    # ── Step 2: Azure Document Intelligence ──────────────────────────────────
    azure_result = {}
    page_texts   = {}
    headings     = []

    if AZURE_DOC_ENDPOINT and AZURE_DOC_KEY:
        azure_result = analyze_pdf_with_azure(pdf_path)
        page_texts   = extract_page_text(azure_result)
        headings     = extract_headings(azure_result)
    else:
        print("  Warning: Azure Doc Intelligence not configured.")

    # Fill missing field names before OpenAI enrichment:
    # Doc Intelligence candidates first, then PdfReader fallback.
    acro_fields = resolve_missing_field_identity(acro_fields, pdf_path)

    # ── Step 3a: Enrich fields via OpenAI ────────────────────────────────────
    if AZURE_OPENAI_ENDPOINT and AZURE_OPENAI_API_KEY and page_texts:
        enriched = enrich_fields_with_openai(acro_fields, page_texts, headings)
    else:
        if not (AZURE_OPENAI_ENDPOINT and AZURE_OPENAI_API_KEY):
            print("  Warning: Azure OpenAI not configured — using field names only.")
        enriched = [
            {
                "page":        f["page"],
                "section":     "Form Fields",
                "field_name":  f["field_name"],
                "label":       f["field_name"],
                "type":        f["field_type"],
                "description": "",
                "tooltip":     "",
            }
            for f in acro_fields
        ]

    fields = assign_sr_nos(enriched)
    return fields, azure_result, page_texts, headings


def generate_data_dictionary(
    input_reports_dir: str = INPUT_REPORTS_DIR,
    output_dir:        str = OUTPUT_DIR,
):
    os.makedirs(output_dir, exist_ok=True)

    if not os.path.isdir(input_reports_dir):
        os.makedirs(input_reports_dir, exist_ok=True)
        print(f"Created input folder: {input_reports_dir}")
        print(f"Place your PDF files there and run again.")
        return

    pdf_files = sorted(
        os.path.join(input_reports_dir, f)
        for f in os.listdir(input_reports_dir)
        if f.lower().endswith(".pdf")
    )
    if not pdf_files:
        print(f"No PDF files found in: {input_reports_dir}")
        return

    print(f"Found {len(pdf_files)} PDF(s) to process.\n")

    for pdf_path in pdf_files:
        pdf_stem  = os.path.splitext(os.path.basename(pdf_path))[0]
        dict_file = os.path.join(output_dir, f"{pdf_stem}_Data_Dictionary.xlsx")
        meta_file = os.path.join(output_dir, f"{pdf_stem}_Metadata.xlsx")

        print(f"Processing: {pdf_stem}.pdf")
        try:
            fields, azure_result, page_texts, headings = process_pdf(pdf_path)
            write_data_dictionary_excel(fields, dict_file, pdf_stem)

            # ── Metadata ──────────────────────────────────────────────────────
            meta = extract_pdf_metadata(pdf_path)

            if AZURE_OPENAI_ENDPOINT and AZURE_OPENAI_API_KEY and page_texts:
                # OpenAI gives accurate Subject, Keywords, Language
                meta = enrich_metadata_with_openai(meta, page_texts, headings)
            elif azure_result:
                # Fallback: derive from layout analysis only
                meta = enrich_metadata_from_azure_only(meta, azure_result)

            write_metadata_excel(meta, meta_file)
            print(f"Done: {pdf_stem}\n")

        except Exception as exc:
            import traceback
            print(f"Error processing {pdf_path}: {exc}")
            traceback.print_exc()
            print()

    print("All done.")


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Generate Data Dictionary and Metadata from PDF form fields."
    )
    parser.add_argument("--input",                default=INPUT_REPORTS_DIR)
    parser.add_argument("--output",               default=OUTPUT_DIR)
    parser.add_argument("--doc-endpoint",         default="", help="Azure Doc Intelligence endpoint")
    parser.add_argument("--doc-key",              default="", help="Azure Doc Intelligence API key")
    parser.add_argument("--openai-endpoint",      default="", help="Azure OpenAI endpoint")
    parser.add_argument("--openai-key",           default="", help="Azure OpenAI API key")
    parser.add_argument("--openai-api-version",   default="", help="Azure OpenAI API version")
    parser.add_argument("--openai-deployment",    default="", help="Azure OpenAI deployment name")
    args = parser.parse_args()

    if args.doc_endpoint:      AZURE_DOC_ENDPOINT       = args.doc_endpoint
    if args.doc_key:           AZURE_DOC_KEY            = args.doc_key
    if args.openai_endpoint:   AZURE_OPENAI_ENDPOINT    = args.openai_endpoint
    if args.openai_key:        AZURE_OPENAI_API_KEY     = args.openai_key
    if args.openai_api_version:AZURE_OPENAI_API_VERSION = args.openai_api_version
    if args.openai_deployment: AZURE_OPENAI_DEPLOYMENT  = args.openai_deployment

    generate_data_dictionary(
        input_reports_dir=args.input,
        output_dir=args.output,
    )
