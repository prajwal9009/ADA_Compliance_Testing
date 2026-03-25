import json
import os
from collections import Counter, defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


INPUT_FOLDER = "Final_reports"
OUTPUT_FOLDER = "Analysis_Output"
OUTPUT_FILE = "WCAG22_AA_Final_Report_Analysis.xlsx"
FINAL_SUMMARY_OUTPUT_FILE = "Final_excel_generated.xlsx"
# Used to provide a "reference to issues file" when you run `Json_Final_Reports_To_Excel.py`.
ISSUES_EXCEL_FOLDER = "Final_report_excels"

EFFORT_SUBCATEGORY_TO_BAND = {
    "simple": "Small",
    "medium": "Medium",
    "complex": "Large",
}


def map_effort_to_fix(effort_subcategory):
    if effort_subcategory is None:
        return ""
    key = str(effort_subcategory).strip().lower()
    return EFFORT_SUBCATEGORY_TO_BAND.get(key, str(effort_subcategory))


def derive_document_stem(document_name, file_name):
    """
    Derive a stable stem like `CCP1000` from either:
    - report["document_name"] (usually `CCP1000.json`)
    - the final report filename (`CCP1000_Final_report.json`)
    """
    if isinstance(document_name, str) and document_name.strip():
        base = os.path.splitext(document_name.strip())[0]
        if base:
            return base

    # Expected: <stem>_Final_report.json
    lowered = (file_name or "").lower()
    if lowered.endswith("_final_report.json"):
        return file_name[: -len("_Final_report.json")]
    return os.path.splitext(file_name)[0]


def resolve_issues_reference(stem):
    """
    Return a relative path to the best available per-document "issues file".
    Prefer per-PDF Excel output when present; otherwise fall back to the per-PDF JSON.
    """
    candidate_excel = os.path.join(ISSUES_EXCEL_FOLDER, f"{stem}.xlsx")
    if os.path.isfile(candidate_excel):
        return candidate_excel
    return os.path.join(INPUT_FOLDER, f"{stem}_Final_report.json")


def safe_number(value, default=0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def normalize_impact(impact_value):
    if not impact_value:
        return "Unknown"
    impact = str(impact_value).strip().lower()
    if impact in ("high", "critical"):
        return "High"
    if impact == "medium":
        return "Medium"
    if impact == "low":
        return "Low"
    return "Unknown"


def calculate_effort_band(unique_failed_issues, high_count, medium_count, low_count, manual_checks, score):
    # Weighted effort score to estimate remediation effort size.
    weighted_score = (high_count * 3) + (medium_count * 2) + low_count + manual_checks

    if unique_failed_issues == 0 and manual_checks == 0 and score >= 95:
        return "Small", weighted_score
    if weighted_score <= 10:
        return "Small", weighted_score
    if weighted_score <= 24:
        return "Medium", weighted_score
    return "Large", weighted_score


def dedupe_issues(issues):
    unique = {}
    for issue in issues:
        key = (
            issue.get("rule", ""),
            issue.get("category", ""),
            issue.get("wcag", ""),
            issue.get("impact", ""),
            issue.get("status", ""),
        )
        if key not in unique:
            unique[key] = issue
    return list(unique.values())


def load_reports(input_folder):
    reports = []
    if not os.path.isdir(input_folder):
        raise FileNotFoundError(f"Input folder not found: {input_folder}")

    files = sorted(
        f for f in os.listdir(input_folder)
        if f.lower().endswith(".json") and os.path.isfile(os.path.join(input_folder, f))
    )
    if not files:
        raise FileNotFoundError(f"No JSON reports found in {input_folder}")

    for file_name in files:
        full_path = os.path.join(input_folder, file_name)
        with open(full_path, "r", encoding="utf-8") as file:
            report = json.load(file)
        reports.append((file_name, report))
    return reports


def auto_fit_columns(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        col_idx = column_cells[0].column
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 12), 55)


def format_sheet(sheet, table_name=None):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.freeze_panes = "A2"
    auto_fit_columns(sheet)

    if table_name and sheet.max_row > 1 and sheet.max_column > 1:
        ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)


def write_executive_summary(workbook, document_rows, generated_on):
    ws = workbook.create_sheet("Executive_Summary")
    ws.append(["Metric", "Value"])

    total_docs = len(document_rows)
    status_counter = Counter(row["overall_status"] for row in document_rows)
    effort_counter = Counter(row["effort_band"] for row in document_rows)

    small_names = [row["pdf_name"] for row in document_rows if row["effort_band"] == "Small"]
    medium_names = [row["pdf_name"] for row in document_rows if row["effort_band"] == "Medium"]
    large_names = [row["pdf_name"] for row in document_rows if row["effort_band"] == "Large"]

    avg_score = round(sum(row["score"] for row in document_rows) / total_docs, 2) if total_docs else 0
    avg_failures = round(sum(row["failed"] for row in document_rows) / total_docs, 2) if total_docs else 0

    rows = [
        ("Generated On (UTC)", generated_on),
        ("Total PDFs Analyzed", total_docs),
        ("Average Compliance Score", avg_score),
        ("Average Failed Checks per PDF", avg_failures),
        ("PASS Count", status_counter.get("PASS", 0)),
        ("FAIL Count", status_counter.get("FAIL", 0)),
        ("REVIEW REQUIRED Count", status_counter.get("REVIEW REQUIRED", 0)),
        ("Small Changes Required - Count", effort_counter.get("Small", 0)),
        ("Small Changes Required - PDFs", ", ".join(small_names) if small_names else "None"),
        ("Medium Changes Required - Count", effort_counter.get("Medium", 0)),
        ("Medium Changes Required - PDFs", ", ".join(medium_names) if medium_names else "None"),
        ("Large Changes Required - Count", effort_counter.get("Large", 0)),
        ("Large Changes Required - PDFs", ", ".join(large_names) if large_names else "None"),
    ]

    for metric, value in rows:
        ws.append([metric, value])

    format_sheet(ws, table_name="ExecutiveSummaryTable")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 85


def write_document_analysis(workbook, document_rows):
    ws = workbook.create_sheet("Document_Analysis")
    ws.append([
        "PDF Name",
        "Overall Status",
        "Compliance Score",
        "Passed Checks",
        "Failed Checks",
        "Manual Checks",
        "Unique Failed Issues",
        "High Impact Issues",
        "Medium Impact Issues",
        "Low Impact Issues",
        "Metadata Missing Fields",
        "Estimated Effort Score",
        "Effort Band (Small/Medium/Large)",
        "Top 5 Failed Rules",
    ])

    for row in document_rows:
        ws.append([
            row["pdf_name"],
            row["overall_status"],
            row["score"],
            row["passed"],
            row["failed"],
            row["manual_check"],
            row["unique_failed_issues"],
            row["high_impact_count"],
            row["medium_impact_count"],
            row["low_impact_count"],
            row["metadata_missing_count"],
            row["effort_score"],
            row["effort_band"],
            row["top_failed_rules"],
        ])

    format_sheet(ws, table_name="DocumentAnalysisTable")

    # Visual score heatmap.
    if ws.max_row > 1:
        ws.conditional_formatting.add(
            f"C2:C{ws.max_row}",
            ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                           mid_type="num", mid_value=50, mid_color="FFEB84",
                           end_type="num", end_value=100, end_color="63BE7B"),
        )


def write_issue_breakdown(workbook, title, rows, headers, table_name):
    ws = workbook.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    format_sheet(ws, table_name=table_name)


def build_analysis():
    reports = load_reports(INPUT_FOLDER)
    document_rows = []
    issue_counter = Counter()
    wcag_counter = Counter()
    category_counter = Counter()
    impact_counter = Counter()
    manual_review_counter = Counter()
    metadata_counter = Counter()

    for file_name, report in reports:
        document_name = report.get("document_name") or ""
        stem = derive_document_stem(document_name, file_name)
        pdf_name = report.get("document_name") or file_name.replace("_Final_report.json", ".pdf")
        pdf_filename = f"{stem}.pdf"
        overall_status = report.get("overall_status", "Unknown")
        score = safe_number(report.get("score", 0))

        summary = report.get("summary", {})
        passed = int(safe_number(summary.get("passed", 0), 0))
        failed = int(safe_number(summary.get("failed", 0), 0))
        manual_check = int(safe_number(summary.get("manual_check", 0), 0))

        issues = report.get("issues", []) or []
        unique_issues = dedupe_issues(issues)
        failed_issues = [i for i in unique_issues if str(i.get("status", "")).lower() == "failed"]
        unique_failed_issues = len(failed_issues)

        per_doc_rule_counter = Counter()
        high_impact_count = 0
        medium_impact_count = 0
        low_impact_count = 0

        for issue in failed_issues:
            rule = issue.get("rule", "Unknown Rule")
            category = issue.get("category", "Unknown Category")
            wcag = issue.get("wcag", "Unknown")
            impact = normalize_impact(issue.get("impact", "Unknown"))

            per_doc_rule_counter[rule] += 1
            issue_counter[(rule, category)] += 1
            wcag_counter[wcag] += 1
            category_counter[category] += 1
            impact_counter[impact] += 1

            if impact == "High":
                high_impact_count += 1
            elif impact == "Medium":
                medium_impact_count += 1
            elif impact == "Low":
                low_impact_count += 1

        manual_review_items = report.get("manual_review_required", []) or []
        for item in manual_review_items:
            manual_review_counter[str(item)] += 1

        metadata_missing_fields = (
            report.get("metadata_validation", {}).get("missing_fields", []) or []
        )
        for field in metadata_missing_fields:
            metadata_counter[str(field)] += 1

        effort_band, effort_score = calculate_effort_band(
            unique_failed_issues=unique_failed_issues,
            high_count=high_impact_count,
            medium_count=medium_impact_count,
            low_count=low_impact_count,
            manual_checks=manual_check,
            score=score,
        )

        top_failed_rules = ", ".join(
            f"{rule} ({count})" for rule, count in per_doc_rule_counter.most_common(5)
        ) or "None"

        document_rows.append({
            "pdf_name": pdf_name,
            "pdf_filename": pdf_filename,
            "overall_status": overall_status,
            "score": score,
            "passed": passed,
            "failed": failed,
            "manual_check": manual_check,
            "unique_failed_issues": unique_failed_issues,
            "high_impact_count": high_impact_count,
            "medium_impact_count": medium_impact_count,
            "low_impact_count": low_impact_count,
            "metadata_missing_count": len(metadata_missing_fields),
            "effort_score": effort_score,
            "effort_band": effort_band,
            "top_failed_rules": top_failed_rules,
            "distributed_category": report.get("distributed_category", ""),
            "effort_to_fix": map_effort_to_fix(report.get("distributed_subcategory", "")),
            "issues_reference": resolve_issues_reference(stem),
        })

    document_rows.sort(key=lambda row: (row["effort_band"], row["score"]))

    workbook = Workbook()
    workbook.remove(workbook.active)
    generated_on = datetime.utcnow().isoformat(timespec="seconds")

    write_executive_summary(workbook, document_rows, generated_on)
    write_document_analysis(workbook, document_rows)

    issue_rows = [
        [rule, category, count]
        for (rule, category), count in issue_counter.most_common()
    ]
    write_issue_breakdown(
        workbook,
        "Issue_Breakdown",
        issue_rows,
        ["Rule", "Category", "Failed Issue Count (Unique)"],
        "IssueBreakdownTable",
    )

    wcag_rows = [[wcag, count] for wcag, count in wcag_counter.most_common()]
    write_issue_breakdown(
        workbook,
        "WCAG_Breakdown",
        wcag_rows,
        ["WCAG Criterion", "Failed Issue Count (Unique)"],
        "WCAGBreakdownTable",
    )

    category_rows = [[category, count] for category, count in category_counter.most_common()]
    write_issue_breakdown(
        workbook,
        "Category_Breakdown",
        category_rows,
        ["Category", "Failed Issue Count (Unique)"],
        "CategoryBreakdownTable",
    )

    impact_rows = [[impact, count] for impact, count in impact_counter.most_common()]
    write_issue_breakdown(
        workbook,
        "Impact_Breakdown",
        impact_rows,
        ["Impact", "Failed Issue Count (Unique)"],
        "ImpactBreakdownTable",
    )

    manual_rows = [[item, count] for item, count in manual_review_counter.most_common()]
    write_issue_breakdown(
        workbook,
        "Manual_Review_Items",
        manual_rows,
        ["Manual Review Item", "Count Across PDFs"],
        "ManualReviewTable",
    )

    metadata_rows = [[field, count] for field, count in metadata_counter.most_common()]
    write_issue_breakdown(
        workbook,
        "Metadata_Gaps",
        metadata_rows,
        ["Metadata Field Missing", "Count Across PDFs"],
        "MetadataGapsTable",
    )

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    output_path = os.path.join(OUTPUT_FOLDER, OUTPUT_FILE)
    workbook.save(output_path)

    # Create the requested "final excel" mapping one row per PDF.
    final_workbook = Workbook()
    final_workbook.remove(final_workbook.active)
    ws_final = final_workbook.create_sheet("Final_Summary")
    ws_final.append(["filename", "categorization", "reference to issues file", "effort to fix"])
    for row in document_rows:
        ws_final.append([
            row.get("pdf_filename", ""),
            row.get("distributed_category", ""),
            row.get("issues_reference", ""),
            row.get("effort_to_fix", ""),
        ])
    format_sheet(ws_final, table_name="FinalSummaryTable")
    final_path = os.path.join(OUTPUT_FOLDER, FINAL_SUMMARY_OUTPUT_FILE)
    final_workbook.save(final_path)
    print(f"Final summary excel generated successfully: {final_path}")

    return output_path


if __name__ == "__main__":
    try:
        file_path = build_analysis()
        print(f"Excel analysis generated successfully: {file_path}")
    except Exception as exc:
        print(f"Failed to generate Excel analysis: {exc}")
